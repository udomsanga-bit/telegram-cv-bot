"""
Telegram CV Collector Bot — with multi-step application form
Steps: Name → Role Applied → Phone Number → Upload CV → Confirmation
"""

import os
import io
import csv
import json
import base64
import logging
import tempfile
import requests
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from dotenv import load_dotenv

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.resolve()
load_dotenv(BASE_DIR / ".env")

BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN not found in .env file!")

# ── Google Drive config ───────────────────────────────────────────────────────
# OAuth2 user credentials (upload as the Google account owner — avoids service
# account quota errors on personal Drive)
GOOGLE_CLIENT_ID     = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.getenv("GOOGLE_REFRESH_TOKEN", "")
GDRIVE_FOLDER_ID     = os.getenv("GOOGLE_DRIVE_FOLDER_ID", "")
GDRIVE_ENABLED       = bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET and
                            GOOGLE_REFRESH_TOKEN and GDRIVE_FOLDER_ID)

# ── Local fallback dirs (used on Railway as temp space) ───────────────────────
CV_DIR = Path(tempfile.gettempdir()) / "collected_cvs"
CV_DIR.mkdir(exist_ok=True)

EXCEL_FILENAME = "submissions.xlsx"
LOG_FILE_CSV   = Path(tempfile.gettempdir()) / "submissions.csv"
LOG_FILE_EXCEL = Path(tempfile.gettempdir()) / EXCEL_FILENAME

HEADERS = [
    "ថ្ងៃខែឆ្នាំ (Timestamp)",
    "Telegram ID",
    "Telegram Username",
    "ឈ្មោះ (Name)",
    "ផ្នែក (Role Applied)",
    "លេខទូរសព្ទ័ (Phone)",
    "ឯកសារ CV (Filename)",
    "ទីតាំងឯកសារ (Saved Path)",
]

ALLOWED_EXTENSIONS = {".pdf", ".doc", ".docx"}
ALLOWED_MIME_TYPES = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(BASE_DIR / "bot.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

# ── In-memory store for form data per user ───────────────────────────────────
# { user_id: { "name": ..., "role": ..., "phone": ... } }
user_data: dict = {}

# ── Clear any webhook so polling works ───────────────────────────────────────
def clear_webhook():
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/deleteWebhook?drop_pending_updates=true"
    try:
        resp = requests.get(url, timeout=10)
        if resp.ok:
            logger.info("Webhook cleared. Polling ready.")
    except Exception as e:
        logger.warning("Could not clear webhook: %s", e)

# ── Google Drive helpers ──────────────────────────────────────────────────────
def _get_drive_service():
    """Build and return an authenticated Google Drive service using OAuth2 user credentials."""
    creds = Credentials(
        token=None,
        refresh_token=GOOGLE_REFRESH_TOKEN,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    creds.refresh(Request())
    return build("drive", "v3", credentials=creds, cache_discovery=False)

def _find_file_in_drive(service, filename: str) -> str | None:
    """Return the Drive file ID of a file by name in the target folder, or None."""
    resp = service.files().list(
        q=f"name='{filename}' and '{GDRIVE_FOLDER_ID}' in parents and trashed=false",
        fields="files(id, name)",
        spaces="drive",
    ).execute()
    files = resp.get("files", [])
    return files[0]["id"] if files else None

def upload_file_to_drive(local_path: Path, filename: str) -> bool:
    """Upload (or overwrite) a file in the shared Google Drive folder."""
    if not GDRIVE_ENABLED:
        return False
    try:
        service   = _get_drive_service()
        mime_type = "application/octet-stream"
        with open(local_path, "rb") as fh:
            media = MediaIoBaseUpload(fh, mimetype=mime_type, resumable=False)
            existing_id = _find_file_in_drive(service, filename)
            if existing_id:
                # Update existing file
                service.files().update(
                    fileId=existing_id, media_body=media
                ).execute()
            else:
                # Create new file in the shared folder
                service.files().create(
                    body={"name": filename, "parents": [GDRIVE_FOLDER_ID]},
                    media_body=media,
                    fields="id",
                ).execute()
        logger.info("Google Drive upload OK → %s", filename)
        return True
    except Exception as e:
        logger.error("Google Drive upload failed: %s", e)
        return False

def download_file_from_drive(filename: str, local_path: Path) -> bool:
    """Download a file from Google Drive to a local path."""
    if not GDRIVE_ENABLED:
        return False
    try:
        service     = _get_drive_service()
        file_id     = _find_file_in_drive(service, filename)
        if not file_id:
            return False   # doesn't exist yet — that's fine
        request     = service.files().get_media(fileId=file_id)
        buf         = io.BytesIO()
        downloader  = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        local_path.write_bytes(buf.getvalue())
        return True
    except Exception as e:
        logger.error("Google Drive download failed: %s", e)
        return False

# ── Bot setup ─────────────────────────────────────────────────────────────────
bot = telebot.TeleBot(BOT_TOKEN, parse_mode=None)

# ── Excel helper ─────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11) # white bold
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")   # light blue alternate rows
COL_WIDTHS   = [22, 14, 20, 24, 18, 18, 36, 50]

def _init_excel():
    """Create the Excel file with a styled header row if it doesn't exist."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submissions"
    ws.append(HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(LOG_FILE_EXCEL)

def _append_excel(row: list):
    """Append one data row to the Excel file with alternating row colour.
    If Google Drive is enabled, pulls the latest version first, then re-uploads.
    """
    # Pull latest from Google Drive so we never overwrite a concurrent update
    if GDRIVE_ENABLED:
        download_file_from_drive(EXCEL_FILENAME, LOG_FILE_EXCEL)

    if not LOG_FILE_EXCEL.exists():
        _init_excel()

    wb = openpyxl.load_workbook(LOG_FILE_EXCEL)
    ws = wb.active
    row_num = ws.max_row + 1
    ws.append(row)
    fill = ALT_FILL if row_num % 2 == 0 else PatternFill()
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    ws.row_dimensions[row_num].height = 18
    wb.save(LOG_FILE_EXCEL)

    # Push updated file back to Google Drive
    if GDRIVE_ENABLED:
        upload_file_to_drive(LOG_FILE_EXCEL, EXCEL_FILENAME)

# ── CSV + Excel logging ───────────────────────────────────────────────────────
def log_submission(telegram_id, telegram_username, name, role, phone, filename, saved_path):
    timestamp = datetime.now().isoformat()
    row = [
        timestamp,
        telegram_id,
        telegram_username or "",
        name,
        role,
        phone,
        filename,
        str(saved_path),
    ]

    # ── Write to Excel ────────────────────────────────────────────────────────
    try:
        _append_excel(row)
    except Exception as e:
        logger.error("Excel log failed: %s", e)

    # ── Write to CSV (backup) ─────────────────────────────────────────────────
    write_header = not LOG_FILE_CSV.exists()
    with open(LOG_FILE_CSV, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if write_header:
            w.writerow(HEADERS)
        w.writerow(row)

# ── Available roles ───────────────────────────────────────────────────────────
# Map button display text → internal role code
ROLE_MAP = {
    "ផ្នែកលក់ស៊ីមកាត":          "FSO",
    "ផ្នែកលក់អុិនធើណិតតាមផ្ទះ": "HISO",
}
ROLE_DISPLAY = {
    "FSO":  "ផ្នែកលក់ស៊ីមកាត (FSO)",
    "HISO": "ផ្នែកលក់អុិនធើណិតតាមផ្ទះ (HISO)",
}

# ── Keyboards ─────────────────────────────────────────────────────────────────
def start_keyboard():
    """Landing button — shown on /start."""
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(KeyboardButton("ដាក់ពាក្យធ្វើការ"))
    return kb

def role_keyboard():
    """Two role buttons with Khmer labels."""
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(KeyboardButton("ផ្នែកលក់ស៊ីមកាត"))
    kb.add(KeyboardButton("ផ្នែកលក់អុិនធើណិតតាមផ្ទះ"))
    return kb

def upload_keyboard():
    """Prompt button reminding the user to attach their CV."""
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(KeyboardButton("📎 ដាក់CV"))
    return kb

def remove_keyboard():
    return ReplyKeyboardRemove()

# ── Step 1: /start → landing page with single Apply button ───────────────────
@bot.message_handler(commands=["start"])
def cmd_start(message):
    user_id = message.from_user.id
    user_data[user_id] = {}  # reset any previous session

    bot.send_message(
        message.chat.id,
        "👋 សូមស្វាគមន៍មកកាន់ Bot ដាក់ពាក្យការងារ!\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🏢 យើងកំពុងស្វែងរកបុគ្គលិកថ្មី!\n\n"
        "ចុចប៊ូតុងខាងក្រោម ដើម្បីចាប់ផ្ដើមដាក់ពាក្យ៖",
        reply_markup=start_keyboard(),
    )
    bot.register_next_step_handler(message, step_landing)

# ── Step 1a: user taps "ដាក់ពាក្យធ្វើការ" → show role selection ──────────────
def step_landing(message):
    user_id = message.from_user.id

    if message.text and message.text.startswith("/"):
        bot.send_message(message.chat.id,
            "⚠️ សូមចុចប៊ូតុង *ដាក់ពាក្យធ្វើការ* ខាងក្រោម ឬផ្ញើ /start ។",
            parse_mode="Markdown",
            reply_markup=start_keyboard())
        bot.register_next_step_handler(message, step_landing)
        return

    if (message.text or "").strip() != "ដាក់ពាក្យធ្វើការ":
        bot.send_message(message.chat.id,
            "👇 សូមចុចប៊ូតុង *ដាក់ពាក្យធ្វើការ* ខាងក្រោម៖",
            parse_mode="Markdown",
            reply_markup=start_keyboard())
        bot.register_next_step_handler(message, step_landing)
        return

    bot.send_message(
        message.chat.id,
        "📋 *ជ្រើសរើសផ្នែកការងារ*\n\n"
        "សូមចុចជ្រើសរើសផ្នែកដែលអ្នកចង់ដាក់ពាក្យ៖",
        parse_mode="Markdown",
        reply_markup=role_keyboard(),
    )
    bot.register_next_step_handler(message, step_select_role)

# ── Step 1b: receive role button → ask for name ───────────────────────────────
def step_select_role(message):
    user_id = message.from_user.id
    choice_text = (message.text or "").strip()
    role = ROLE_MAP.get(choice_text)

    if message.text and message.text.startswith("/"):
        bot.send_message(message.chat.id,
            "⚠️ សូមចុចជ្រើសរើសផ្នែក ឬផ្ញើ /start ដើម្បីចាប់ផ្ដើមឡើងវិញ។",
            reply_markup=role_keyboard())
        bot.register_next_step_handler(message, step_select_role)
        return

    if role is None:
        bot.send_message(
            message.chat.id,
            "⚠️ សូមចុចប៊ូតុងខាងក្រោម ដើម្បីជ្រើសរើសផ្នែក៖",
            reply_markup=role_keyboard(),
        )
        bot.register_next_step_handler(message, step_select_role)
        return

    user_data[user_id]["role"] = role

    bot.send_message(
        message.chat.id,
        f"✅ បានជ្រើសរើស *{ROLE_DISPLAY[role]}*!\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📝 *ជំហានទី 1/3 — ឈ្មោះពេញ*\n\n"
        "វាយបញ្ចូលឈ្មោះពេញរបស់អ្នក៖",
        parse_mode="Markdown",
        reply_markup=remove_keyboard(),
    )
    bot.register_next_step_handler(message, step_get_name)

# ── Step 2: receive name → ask for phone ─────────────────────────────────────
def step_get_name(message):
    user_id = message.from_user.id

    if message.text and message.text.startswith("/"):
        bot.send_message(message.chat.id, "⚠️ សូមវាយបញ្ចូលឈ្មោះពេញរបស់អ្នក ឬផ្ញើ /start ដើម្បីចាប់ផ្ដើមឡើងវិញ។")
        bot.register_next_step_handler(message, step_get_name)
        return

    name = (message.text or "").strip()
    if len(name) < 2:
        bot.send_message(message.chat.id, "⚠️ ឈ្មោះខ្លីពេក។ សូមវាយបញ្ចូលឈ្មោះពេញរបស់អ្នក៖")
        bot.register_next_step_handler(message, step_get_name)
        return

    user_data[user_id]["name"] = name

    bot.send_message(
        message.chat.id,
        f"✅ បានទទួល *{name}*!\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📞 *ជំហានទី 2/3 — លេខទូរសព្ទ័*\n\n"
        "សូមបញ្ចូលលេខទូរសព្ទ័របស់អ្នក (រួមទាំងលេខប្រទេស ឧ. +85512345678):",
        parse_mode="Markdown",
    )
    bot.register_next_step_handler(message, step_get_phone)

# ── Step 4: receive phone → ask for CV ───────────────────────────────────────
def step_get_phone(message):
    user_id = message.from_user.id

    if message.text and message.text.startswith("/"):
        bot.send_message(message.chat.id, "⚠️ សូមវាយបញ្ចូលលេខទូរសព្ទ័ ឬផ្ញើ /start ដើម្បីចាប់ផ្ដើមឡើងវិញ។")
        bot.register_next_step_handler(message, step_get_phone)
        return

    phone = (message.text or "").strip()
    # Basic validation: must contain at least 7 digits
    digits = [c for c in phone if c.isdigit()]
    if len(digits) < 7:
        bot.send_message(
            message.chat.id,
            "⚠️ លេខទូរសព្ទ័មិនត្រឹមត្រូវ។\n"
            "សូមបញ្ចូលលេខរួមទាំងលេខប្រទេស (ឧ. +85512345678):"
        )
        bot.register_next_step_handler(message, step_get_phone)
        return

    user_data[user_id]["phone"] = phone

    name = user_data[user_id].get("name", "")
    bot.send_message(
        message.chat.id,
        f"✅ បានទទួលលេខទូរសព្ទ័ *{phone}*!\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📄 *ជំហានទី 3/3 — ដាក់CV*\n\n"
        f"ជិតដល់ហើយ {name}! 🎉\n\n"
        "សូមភ្ជាប់ និងផ្ញើឯកសារ CV របស់អ្នក។\n"
        "ទម្រង់ដែលទទួលយក៖ *PDF, DOC, DOCX*\n\n"
        "👇 ចុចប៊ូតុង 📎 ខាងក្រោម ជ្រើស *ឯកសារ (File)* រួចផ្ញើ CV របស់អ្នក។",
        parse_mode="Markdown",
        reply_markup=upload_keyboard(),
    )
    bot.register_next_step_handler(message, step_get_cv)

# ── Step 5: receive CV file → save + confirm ─────────────────────────────────
def step_get_cv(message):
    user_id = message.from_user.id

    # User tapped the reminder button or sent text — keep waiting for the file
    if not message.document:
        if message.photo:
            bot.send_message(
                message.chat.id,
                "📸 នេះជារូបភាព មិនមែនឯកសារទេ។\n\n"
                "សូមភ្ជាប់ CV ជា *ឯកសារ (Document)* (PDF/DOC/DOCX)។\n"
                "ចុច 📎 → ជ្រើស *ឯកសារ (File)* មិនមែនរូបភាព។",
                parse_mode="Markdown",
                reply_markup=upload_keyboard(),
            )
        else:
            bot.send_message(
                message.chat.id,
                "👆 សូមភ្ជាប់ឯកសារ CV (PDF, DOC, ឬ DOCX) រួចផ្ញើ។",
                reply_markup=upload_keyboard(),
            )
        bot.register_next_step_handler(message, step_get_cv)
        return

    doc = message.document

    # Validate file type
    ext = Path(doc.file_name or "").suffix.lower()
    mime_ok = (doc.mime_type or "") in ALLOWED_MIME_TYPES
    ext_ok = ext in ALLOWED_EXTENSIONS

    if not (mime_ok or ext_ok):
        bot.send_message(
            message.chat.id,
            f"⚠️ '{doc.file_name}' មិនមែនជាទម្រង់ដែលទទួលយកទេ។\n"
            "សូមផ្ទុកឡើងឯកសារ PDF, DOC, ឬ DOCX។",
            reply_markup=upload_keyboard(),
        )
        bot.register_next_step_handler(message, step_get_cv)
        return

    if doc.file_size and doc.file_size > 20 * 1024 * 1024:
        bot.send_message(
            message.chat.id,
            "⚠️ ឯកសារធំពេក (អតិបរមា 20 MB)។ សូមបង្ហាប់ហើយព្យាយាមម្ដងទៀត។",
            reply_markup=upload_keyboard(),
        )
        bot.register_next_step_handler(message, step_get_cv)
        return

    # Save file
    data = user_data.get(user_id, {})
    name      = data.get("name", "unknown")
    role      = data.get("role", "unknown")
    role_disp = ROLE_DISPLAY.get(role, role)
    phone     = data.get("phone", "unknown")

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_uname = (message.from_user.username or f"user{user_id}").replace(" ", "_")
    original   = doc.file_name or f"cv_{timestamp}{ext or '.pdf'}"
    save_name  = f"{timestamp}_{safe_uname}_{original}"
    save_path  = CV_DIR / save_name

    bot.send_message(message.chat.id, "⏳ កំពុងរក្សាទុក CV របស់អ្នក…", reply_markup=remove_keyboard())

    try:
        file_info  = bot.get_file(doc.file_id)
        downloaded = bot.download_file(file_info.file_path)
        with open(save_path, "wb") as f:
            f.write(downloaded)
    except Exception as e:
        logger.error("Download failed for user %s: %s", user_id, e)
        bot.send_message(
            message.chat.id,
            "❌ មានបញ្ហាក្នុងការរក្សាទុកឯកសាររបស់អ្នក។ សូមព្យាយាមម្ដងទៀត។",
            reply_markup=upload_keyboard(),
        )
        bot.register_next_step_handler(message, step_get_cv)
        return

    # Upload CV to Google Drive
    gdrive_ok = upload_file_to_drive(save_path, save_name)
    storage_note = (
        f"☁️ Google Drive: {save_name}" if gdrive_ok else f"💾 Local: {save_path}"
    )

    log_submission(
        telegram_id=user_id,
        telegram_username=message.from_user.username,
        name=name,
        role=role,
        phone=phone,
        filename=original,
        saved_path=f"Google Drive/{save_name}" if gdrive_ok else str(save_path),
    )

    logger.info("Application saved — %s | %s | %s | %s", name, role, phone, storage_note)

    # Clean up stored data
    user_data.pop(user_id, None)

    # ── Confirmation message ──────────────────────────────────────────────────
    bot.send_message(
        message.chat.id,
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🎉 *បានទទួលពាក្យសុំការងាររបស់អ្នក!*\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "អរគុណចំពោះការដាក់ពាក្យ! នេះជាសង្ខេបនៃព័ត៌មានដែលយើងបានទទួល៖\n\n"
        f"👤 *ឈ្មោះ៖* {name}\n"
        f"💼 *ផ្នែក៖* {role_disp}\n"
        f"📞 *លេខទូរសព្ទ័៖* {phone}\n"
        f"📄 *CV៖* {original}\n\n"
        "ក្រុមការងាររបស់យើងនឹងពិនិត្យពាក្យសុំរបស់អ្នក ហើយនឹងទំនាក់ទំនងមកវិញឆាប់ៗ។\n\n"
        "សូមជូនពរជោគជ័យ! 🍀\n\n"
        "_ផ្ញើ /start ប្រសិនបើអ្នកចង់ដាក់ពាក្យម្ដងទៀត។_",
        parse_mode="Markdown",
        reply_markup=remove_keyboard(),
    )

# ── Fallback for messages outside the form flow ──────────────────────────────
@bot.message_handler(func=lambda m: True)
def handle_fallback(message):
    bot.send_message(
        message.chat.id,
        "👋 ផ្ញើ /start ដើម្បីចាប់ផ្ដើមដាក់ពាក្យការងារ។",
        reply_markup=remove_keyboard(),
    )

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    clear_webhook()
    me = bot.get_me()
    logger.info("Bot online: @%s", me.username)
    logger.info("CVs will be saved to: %s", CV_DIR)
    logger.info("Press Ctrl+C to stop.\n")
    bot.infinity_polling(timeout=20, long_polling_timeout=20)
